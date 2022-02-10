import PIL.Image
from openpyxl import Workbook
import openpyxl
import os

# Resize Image
def resize_image(image, new_width=100):
    width, height = image.size
    ratio = height / width
    new_height = int(new_width * ratio)
    resized_image = image.resize((new_width, new_height), PIL.Image.NEAREST)
    return resized_image

def pixel_to_Excelpoint(pixel:int):
    return pixel / 7

def pixel_to_excel(image, margenX=0, margenY=0):
    workbook = Workbook()
    sheet = workbook.active

    # Resize the excel sheet
    WIDTH = pixel_to_Excelpoint(20)
    X_LENGTH_PIXEL = image.size[0];

    # # Format every column with width 20px
    for col in range(1, sheet.max_column + X_LENGTH_PIXEL + margenX):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = WIDTH

    # Get the hexadecimal value of each pixel and fill the cells with it
    for i in range(0, image.size[0]):
        for j in range(0, image.size[1]):
            pixel = image.getpixel((i, j))
            # Convert RGB into hexadecimal
            hex_color = '%02x%02x%02x' % (pixel[0], pixel[1], pixel[2])
            # Fill the cell with the hexadecimal value
            sheet.cell(row=j+margenY+1, column=i+margenX+1).fill = openpyxl.styles.PatternFill(patternType="solid", fgColor=hex_color)

    workbook.save("draw.xlsx")


def draw_image(image_path:str, image_size: int)->None:
    image = PIL.Image.open(image_path)
    image = resize_image(image, image_size)
    pixel_to_excel(image)

def main()->None:
    # Ask user for the image path and size
    image_path = input("Enter the image path: ")

    # Check if the image exists
    if not os.path.isfile(image_path):
        print("The image path is not valid :(")
        return

    image_size = int(input("Enter the image size to have in Excel: "))

    # Avoid negative values and higher than 1000
    if image_size <= 0 or image_size > 1000:
        print("The image size is not valid :(")
        return

    draw_image(image_path, image_size)
    print("Image drawn successfully :)")

if __name__ == '__main__':
    main()